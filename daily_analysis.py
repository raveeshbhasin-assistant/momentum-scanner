"""
Daily Analysis Automation
─────────────────────────
Fetches scanner picks, resolves WIN/LOSS/EOD using Yahoo Finance 5-min data,
and appends results to scanner_trade_log.xlsx.

Usage:
    python daily_analysis.py                  # Analyze today
    python daily_analysis.py 2026-04-17       # Analyze a specific date
    python daily_analysis.py --from-json data/2026-04-17.json   # Use local JSON file

Requirements:
    pip install openpyxl httpx --break-system-packages
"""

import argparse
import json
import logging
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

import httpx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers

from performance_engine import (
    assign_category,
    normalize_entry,
    upsert_day,
)

# ── Config ──────────────────────────────────────────────────────
ET = ZoneInfo("America/New_York")
SCANNER_URL = "https://momentum-scanner-production-20b1.up.railway.app"
YAHOO_URL = "https://query1.finance.yahoo.com/v8/finance/chart"
XLSX_PATH = Path(__file__).parent / "scanner_trade_log.xlsx"
DATA_DIR = Path(__file__).parent / "data"

MARKET_OPEN_HOUR, MARKET_OPEN_MIN = 9, 30
MARKET_CLOSE_HOUR, MARKET_CLOSE_MIN = 16, 0

# Known stale-ticker remaps applied before Yahoo fetch.
# SQ (Block, Inc.) rebranded to XYZ in Jan 2025 — Yahoo returns 404 for SQ.
STALE_TICKER_MAP = {
    "SQ": "XYZ",
}

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

# ── Excel styling ───────────────────────────────────────────────
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GREY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
GREEN_FONT = Font(color="006100")
RED_FONT = Font(color="9C0006")
GREY_FONT = Font(color="808080")
MONO_FONT = Font(name="Consolas", size=10)
PCT_FMT = "0.00%"


# ═══════════════════════════════════════════════════════════════
#  STEP 1: FETCH PICKS
# ═══════════════════════════════════════════════════════════════

def fetch_picks_from_scanner(date_str: str) -> list[dict]:
    """Fetch picks from the live scanner's /api/today or daily JSON."""
    # Try local JSON first (data/ directory from the running scanner)
    local_path = DATA_DIR / f"{date_str}.json"
    if local_path.exists():
        logger.info(f"Loading picks from local file: {local_path}")
        with open(local_path) as f:
            return json.load(f)

    # Fall back to the live scanner API
    logger.info(f"Fetching picks from scanner API...")
    try:
        with httpx.Client(timeout=30) as client:
            resp = client.get(f"{SCANNER_URL}/api/today")
            resp.raise_for_status()
            data = resp.json()
            return data.get("finds", [])
    except Exception as e:
        logger.error(f"Failed to fetch from scanner: {e}")
        return []


def fetch_picks_from_json(json_path: str) -> list[dict]:
    """Load picks from an explicit JSON file path."""
    path = Path(json_path)
    if not path.exists():
        logger.error(f"JSON file not found: {path}")
        return []
    with open(path) as f:
        return json.load(f)


def deduplicate_picks(picks: list[dict]) -> list[dict]:
    """
    Deduplicate picks by (ticker, entry_price) — keep first appearance.
    Track appearance number for repeat entries at different prices.
    """
    seen = {}  # ticker -> list of entry prices seen
    deduped = []

    for pick in picks:
        ticker = pick["ticker"]
        entry = round(pick.get("entry", pick.get("price", 0)), 2)

        if ticker not in seen:
            seen[ticker] = []

        # Check if this exact entry price was already seen
        if entry in seen[ticker]:
            continue  # Skip duplicate

        seen[ticker].append(entry)
        appearance = len(seen[ticker])

        deduped.append({
            **pick,
            "entry_price": entry,
            "appearance_num": appearance,
        })

    logger.info(f"Deduplicated: {len(picks)} raw picks → {len(deduped)} unique trades")
    return deduped


def filter_market_hours(picks: list[dict]) -> list[dict]:
    """
    Keep picks with found_time ∈ [09:30, 16:00) ET.

    Picks before 9:30 are pre-market noise. Picks at/after 16:00 are post-close
    scanner emissions (a known data-quality bug — APScheduler should gate at
    15:55 ET) and are not actionable as intraday trades, so we drop them.
    """
    kept = []
    dropped_pre = 0
    dropped_post = 0
    for pick in picks:
        time_str = pick.get("found_time", "")
        try:
            t = datetime.strptime(time_str.replace(" ET", "").strip(), "%I:%M %p")
            total_min = t.hour * 60 + t.minute
            market_open_min = MARKET_OPEN_HOUR * 60 + MARKET_OPEN_MIN
            market_close_min = MARKET_CLOSE_HOUR * 60 + MARKET_CLOSE_MIN
            if total_min < market_open_min:
                dropped_pre += 1
                continue
            if total_min >= market_close_min:
                dropped_post += 1
                continue
            kept.append(pick)
        except (ValueError, AttributeError):
            # Can't parse time — include it (conservative)
            kept.append(pick)

    logger.info(
        f"Market hours filter: {len(picks)} → {len(kept)} picks "
        f"(dropped {dropped_pre} pre-open, {dropped_post} post-close)"
    )
    return kept


def _extract_leadership_label(pick: dict) -> str | None:
    """
    Pull the leadership label (LEADER / FOLLOWER / LAGGARD / SOLO_MOVER / UNKNOWN)
    off a raw pick. Scanner stores it under 'leadership' as a dict or sometimes
    directly on the signal.
    """
    lead = pick.get("leadership")
    if isinstance(lead, dict):
        lbl = lead.get("label")
        if lbl:
            return str(lbl).upper()
    if isinstance(lead, str) and lead:
        return lead.upper()
    lbl = pick.get("leadership_label")
    if lbl:
        return str(lbl).upper()
    return None


def _batch_time_24h(found_time: str) -> str:
    """'09:31 AM ET' → '09:31'. Falls back to the raw string if it can't parse."""
    try:
        t = datetime.strptime(found_time.replace(" ET", "").strip(), "%I:%M %p")
        return t.strftime("%H:%M")
    except (ValueError, AttributeError):
        return found_time or ""


# ═══════════════════════════════════════════════════════════════
#  STEP 2: FETCH YAHOO FINANCE DATA
# ═══════════════════════════════════════════════════════════════

def fetch_yahoo_intraday(ticker: str, client: httpx.Client) -> list[dict]:
    """
    Fetch 5-minute intraday OHLC for a ticker from Yahoo Finance.
    Returns list of {timestamp, open, high, low, close, volume} dicts.
    """
    # Handle BRK-B → BRK-B (Yahoo uses hyphens)
    yahoo_ticker = STALE_TICKER_MAP.get(ticker, ticker)
    if yahoo_ticker != ticker:
        logger.info(f"Remapping stale ticker {ticker} → {yahoo_ticker} for Yahoo fetch")

    try:
        resp = client.get(
            f"{YAHOO_URL}/{yahoo_ticker}",
            params={"interval": "5m", "range": "1d"},
            headers={"User-Agent": "Mozilla/5.0"},
        )
        resp.raise_for_status()
        data = resp.json()

        chart = data.get("chart", {}).get("result", [{}])[0]
        timestamps = chart.get("timestamp", [])
        quote = chart.get("indicators", {}).get("quote", [{}])[0]

        bars = []
        for i, ts in enumerate(timestamps):
            bars.append({
                "timestamp": ts,
                "datetime": datetime.fromtimestamp(ts, tz=ET),
                "open": quote.get("open", [None])[i],
                "high": quote.get("high", [None])[i],
                "low": quote.get("low", [None])[i],
                "close": quote.get("close", [None])[i],
                "volume": quote.get("volume", [None])[i],
            })
        return bars

    except Exception as e:
        logger.warning(f"Yahoo Finance fetch failed for {ticker}: {e}")
        return []


# ═══════════════════════════════════════════════════════════════
#  STEP 3: RESOLVE WIN/LOSS/EOD
# ═══════════════════════════════════════════════════════════════

def parse_pick_time(pick: dict, date_str: str) -> datetime | None:
    """Parse the found_time from a pick into a datetime."""
    time_str = pick.get("found_time", "")
    try:
        t = datetime.strptime(time_str.replace(" ET", "").strip(), "%I:%M %p")
        d = datetime.strptime(date_str, "%Y-%m-%d")
        return datetime(d.year, d.month, d.day, t.hour, t.minute, tzinfo=ET)
    except (ValueError, AttributeError):
        return None


def resolve_trade(pick: dict, bars: list[dict], date_str: str) -> dict:
    """
    Determine WIN/LOSS/EOD for a single pick given Yahoo 5-min bars.

    Rules:
    - Scan bars from signal time forward
    - First bar where high >= target → WIN (exit at target)
    - First bar where low <= stop → LOSS (exit at stop)
    - If both in same bar → assume STOP hit first (conservative)
    - If neither by 4 PM → EOD (exit at close of last bar)
    """
    entry_price = pick["entry_price"]
    target = round(pick.get("atr_target", pick.get("target", 0)), 2)
    stop = round(pick.get("stop_loss", pick.get("stop", 0)), 2)

    signal_time = parse_pick_time(pick, date_str)
    if signal_time is None:
        return {"result": "EOD", "exit_price": entry_price, "pnl_dollar": 0, "pnl_pct": 0}

    market_close = signal_time.replace(hour=MARKET_CLOSE_HOUR, minute=MARKET_CLOSE_MIN)
    last_close_price = entry_price  # fallback

    for bar in bars:
        bar_time = bar["datetime"]

        # Skip bars before signal time
        if bar_time < signal_time:
            continue

        # Skip bars after market close
        if bar_time >= market_close:
            break

        high = bar.get("high")
        low = bar.get("low")
        close = bar.get("close")

        if high is None or low is None or close is None:
            continue

        last_close_price = close

        # Check if both target and stop hit in same bar — stop wins (conservative)
        stop_hit = low <= stop if stop > 0 else False
        target_hit = high >= target if target > 0 else False

        if stop_hit and target_hit:
            # Ambiguous bar — assume stop hit first
            exit_price = stop
            pnl = round(exit_price - entry_price, 2)
            pnl_pct = round(pnl / entry_price, 5) if entry_price else 0
            return {"result": "LOSS", "exit_price": exit_price, "pnl_dollar": pnl, "pnl_pct": pnl_pct}

        if stop_hit:
            exit_price = stop
            pnl = round(exit_price - entry_price, 2)
            pnl_pct = round(pnl / entry_price, 5) if entry_price else 0
            return {"result": "LOSS", "exit_price": exit_price, "pnl_dollar": pnl, "pnl_pct": pnl_pct}

        if target_hit:
            exit_price = target
            pnl = round(exit_price - entry_price, 2)
            pnl_pct = round(pnl / entry_price, 5) if entry_price else 0
            return {"result": "WIN", "exit_price": exit_price, "pnl_dollar": pnl, "pnl_pct": pnl_pct}

    # Neither triggered → EOD
    exit_price = round(last_close_price, 2)
    pnl = round(exit_price - entry_price, 2)
    pnl_pct = round(pnl / entry_price, 5) if entry_price else 0
    return {"result": "EOD", "exit_price": exit_price, "pnl_dollar": pnl, "pnl_pct": pnl_pct}


# ═══════════════════════════════════════════════════════════════
#  STEP 4: WRITE TO EXCEL
# ═══════════════════════════════════════════════════════════════

def format_batch_time(found_time: str) -> str:
    """Convert '09:31 AM ET' → '09:31' for the Batch Time column."""
    try:
        t = datetime.strptime(found_time.replace(" ET", "").strip(), "%I:%M %p")
        return t.strftime("%H:%M")
    except (ValueError, AttributeError):
        return found_time


def append_to_excel(trades: list[dict], date_str: str):
    """Append resolved trades to the Trade Log sheet and update summary sheets."""
    if not XLSX_PATH.exists():
        logger.error(f"Excel file not found: {XLSX_PATH}")
        return

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws_log = wb["Trade Log"]
    ws_summary = wb["Daily Summary"]
    ws_batch = wb["Batch Analysis"]

    # Check if this date already has data
    existing_dates = set()
    for row in ws_log.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            existing_dates.add(str(row[0]))

    if date_str in existing_dates:
        logger.warning(f"Date {date_str} already exists in Trade Log — skipping to avoid duplicates")
        logger.info("To re-analyze, first remove existing rows for this date from the Excel")
        wb.close()
        return

    # Append trade rows
    start_row = ws_log.max_row + 1
    for i, trade in enumerate(trades):
        row_num = start_row + i
        batch_time = format_batch_time(trade.get("found_time", ""))
        result = trade["resolution"]["result"]

        row_data = [
            date_str,                                   # Date
            batch_time,                                 # Batch Time (ET)
            trade["ticker"],                            # Ticker
            trade["entry_price"],                       # Entry Price
            round(trade.get("atr_target", trade.get("target", 0)), 2),  # Target
            round(trade.get("stop_loss", trade.get("stop", 0)), 2),     # Stop
            round(trade.get("rvol", 0), 1),             # RVOL
            round(trade.get("rsi", 0) or 0, 0),        # RSI
            round(trade.get("composite_score", 0), 0),  # Score
            result,                                     # Result
            trade["resolution"]["exit_price"],           # Exit Price
            trade["resolution"]["pnl_dollar"],           # P&L $
            trade["resolution"]["pnl_pct"],              # P&L %
            trade.get("appearance_num", 1),              # Appearance #
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws_log.cell(row=row_num, column=col, value=val)
            cell.font = MONO_FONT

            # Format P&L % column
            if col == 13:
                cell.number_format = PCT_FMT

        # Color-code the Result cell
        result_cell = ws_log.cell(row=row_num, column=10)
        if result == "WIN":
            result_cell.fill = GREEN_FILL
            result_cell.font = Font(name="Consolas", size=10, color="006100", bold=True)
        elif result == "LOSS":
            result_cell.fill = RED_FILL
            result_cell.font = Font(name="Consolas", size=10, color="9C0006", bold=True)
        else:
            result_cell.fill = GREY_FILL
            result_cell.font = Font(name="Consolas", size=10, color="808080")

    logger.info(f"Appended {len(trades)} trade rows to Trade Log (rows {start_row}–{start_row + len(trades) - 1})")

    # ── Update Daily Summary ────────────────────────────────
    # Find the TOTAL row and insert a new day row before it
    total_row = None
    for row in ws_summary.iter_rows(min_row=1, max_col=1, values_only=False):
        if row[0].value == "TOTAL":
            total_row = row[0].row
            break

    if total_row:
        # Insert new row before TOTAL
        ws_summary.insert_rows(total_row)
        new_row = total_row  # the inserted row takes this position
        d = f'"{date_str}"'

        ws_summary.cell(row=new_row, column=1, value=date_str)
        ws_summary.cell(row=new_row, column=2, value=f"=COUNTIF('Trade Log'!$A:$A,{d})")
        ws_summary.cell(row=new_row, column=3, value=f'=COUNTIFS(\'Trade Log\'!$A:$A,{d},\'Trade Log\'!$J:$J,"WIN")')
        ws_summary.cell(row=new_row, column=4, value=f'=COUNTIFS(\'Trade Log\'!$A:$A,{d},\'Trade Log\'!$J:$J,"LOSS")')
        ws_summary.cell(row=new_row, column=5, value=f'=COUNTIFS(\'Trade Log\'!$A:$A,{d},\'Trade Log\'!$J:$J,"EOD")')
        ws_summary.cell(row=new_row, column=6, value=f"=IFERROR(C{new_row}/(C{new_row}+D{new_row}),\"-\")")
        ws_summary.cell(row=new_row, column=7, value=f"=SUMIF('Trade Log'!$A:$A,{d},'Trade Log'!$L:$L)")
        ws_summary.cell(row=new_row, column=8, value=f'=IFERROR(AVERAGEIFS(\'Trade Log\'!$L:$L,\'Trade Log\'!$A:$A,{d},\'Trade Log\'!$J:$J,"WIN"),"-")')
        ws_summary.cell(row=new_row, column=9, value=f'=IFERROR(ABS(AVERAGEIFS(\'Trade Log\'!$L:$L,\'Trade Log\'!$A:$A,{d},\'Trade Log\'!$J:$J,"LOSS")),"-")')
        ws_summary.cell(row=new_row, column=10, value=f"=IFERROR(H{new_row}/I{new_row},\"-\")")

        # Update TOTAL row formulas (now shifted down by 1)
        total_row_new = total_row + 1
        last_data_row = total_row_new - 1
        ws_summary.cell(row=total_row_new, column=1, value="TOTAL")
        ws_summary.cell(row=total_row_new, column=2, value=f"=SUM(B2:B{last_data_row})")
        ws_summary.cell(row=total_row_new, column=3, value=f"=SUM(C2:C{last_data_row})")
        ws_summary.cell(row=total_row_new, column=4, value=f"=SUM(D2:D{last_data_row})")
        ws_summary.cell(row=total_row_new, column=5, value=f"=SUM(E2:E{last_data_row})")
        ws_summary.cell(row=total_row_new, column=6, value=f"=IFERROR(C{total_row_new}/(C{total_row_new}+D{total_row_new}),\"-\")")
        ws_summary.cell(row=total_row_new, column=7, value=f"=SUM(G2:G{last_data_row})")

        logger.info(f"Added Daily Summary row for {date_str}")

    # ── Update Batch Analysis ───────────────────────────────
    # Get unique batch times for this date
    batch_times = sorted(set(format_batch_time(t.get("found_time", "")) for t in trades))
    batch_start_row = ws_batch.max_row + 1

    for i, bt in enumerate(batch_times):
        row_num = batch_start_row + i
        d = f'"{date_str}"'
        b = f'"{bt}"'

        ws_batch.cell(row=row_num, column=1, value=date_str)
        ws_batch.cell(row=row_num, column=2, value=bt)
        ws_batch.cell(row=row_num, column=3, value=f"=COUNTIFS('Trade Log'!$A:$A,{d},'Trade Log'!$B:$B,{b})")
        ws_batch.cell(row=row_num, column=4, value=f'=COUNTIFS(\'Trade Log\'!$A:$A,{d},\'Trade Log\'!$B:$B,{b},\'Trade Log\'!$J:$J,"WIN")')
        ws_batch.cell(row=row_num, column=5, value=f'=COUNTIFS(\'Trade Log\'!$A:$A,{d},\'Trade Log\'!$B:$B,{b},\'Trade Log\'!$J:$J,"LOSS")')
        ws_batch.cell(row=row_num, column=6, value=f'=COUNTIFS(\'Trade Log\'!$A:$A,{d},\'Trade Log\'!$B:$B,{b},\'Trade Log\'!$J:$J,"EOD")')
        ws_batch.cell(row=row_num, column=7, value=f"=IFERROR(D{row_num}/(D{row_num}+E{row_num}),\"-\")")

    logger.info(f"Added {len(batch_times)} Batch Analysis rows for {date_str}")

    # Save
    wb.save(XLSX_PATH)
    wb.close()
    logger.info(f"Excel saved: {XLSX_PATH}")


# ═══════════════════════════════════════════════════════════════
#  MAIN WORKFLOW
# ═══════════════════════════════════════════════════════════════

def analyze_day(date_str: str, json_path: str = None):
    """Full daily analysis workflow."""
    logger.info(f"═══ Daily Analysis for {date_str} ═══")

    # Step 1: Get picks
    if json_path:
        raw_picks = fetch_picks_from_json(json_path)
    else:
        raw_picks = fetch_picks_from_scanner(date_str)

    if not raw_picks:
        logger.error("No picks found — nothing to analyze")
        return

    # Step 2: Filter and deduplicate
    picks = filter_market_hours(raw_picks)
    picks = deduplicate_picks(picks)

    if not picks:
        logger.error("No picks after filtering — nothing to analyze")
        return

    # Step 3: Fetch Yahoo Finance data and resolve each trade
    unique_tickers = list(set(p["ticker"] for p in picks))
    logger.info(f"Fetching Yahoo Finance 5-min data for {len(unique_tickers)} tickers...")

    yahoo_data = {}
    with httpx.Client(timeout=15) as client:
        for i, ticker in enumerate(unique_tickers):
            bars = fetch_yahoo_intraday(ticker, client)
            if bars:
                yahoo_data[ticker] = bars
                logger.info(f"  {ticker}: {len(bars)} bars")
            else:
                logger.warning(f"  {ticker}: no data")

            # Rate limit — Yahoo is generous but let's be polite
            if (i + 1) % 5 == 0:
                time.sleep(1.0)

    logger.info(f"Yahoo data fetched for {len(yahoo_data)}/{len(unique_tickers)} tickers")

    # Step 4: Resolve WIN/LOSS/EOD for each pick
    wins, losses, eods = 0, 0, 0
    total_pnl = 0.0

    for pick in picks:
        bars = yahoo_data.get(pick["ticker"], [])
        resolution = resolve_trade(pick, bars, date_str)
        pick["resolution"] = resolution

        if resolution["result"] == "WIN":
            wins += 1
        elif resolution["result"] == "LOSS":
            losses += 1
        else:
            eods += 1
        total_pnl += resolution["pnl_dollar"]

    decided = wins + losses
    win_rate = (wins / decided * 100) if decided > 0 else 0

    logger.info(f"")
    logger.info(f"═══ Results for {date_str} ═══")
    logger.info(f"  Trades: {len(picks)} ({wins}W / {losses}L / {eods}EOD)")
    logger.info(f"  Win rate (decided): {win_rate:.1f}% ({wins}/{decided})")
    logger.info(f"  Net P&L: ${total_pnl:+.2f}")
    logger.info(f"")

    # Step 5: Write to Excel
    append_to_excel(picks, date_str)

    # Step 6: Write to the v3.5 performance log (drives /performance)
    try:
        _write_performance_log(picks, date_str)
    except Exception as e:
        logger.error(f"Failed to write performance_log.json: {e}")

    # Step 7: Mirror the critical data files to GitHub (belt-and-suspenders
    # backup of the Railway volume). No-ops cleanly if env vars aren't set.
    try:
        from data_backup import backup_data_files
        backup_data_files(
            files=[DATA_DIR / "performance_log.json",
                   DATA_DIR / f"{date_str}.json"],
            tag=f"EOD {date_str}",
        )
    except Exception as e:
        logger.warning(f"Data backup step failed (non-fatal): {e}")

    # Print summary
    print(f"\n{'='*50}")
    print(f"  {date_str} — {len(picks)} trades analyzed")
    print(f"  {wins}W / {losses}L / {eods}EOD")
    print(f"  Win rate: {win_rate:.1f}%")
    print(f"  Net P&L: ${total_pnl:+.2f}")
    print(f"  → Appended to {XLSX_PATH.name}")
    print(f"{'='*50}\n")


def _write_performance_log(picks: list[dict], date_str: str):
    """
    Project resolved picks into the v3.5 performance log schema and upsert
    the day. Idempotent — re-runs for the same date replace that day's rows.
    """
    entries = []
    cat_counts = {"A": 0, "B": 0, "C": 0, "D": 0}
    for p in picks:
        res = p.get("resolution") or {}
        label = _extract_leadership_label(p)
        score = p.get("composite_score", p.get("score", 0)) or 0
        entry_price = p.get("entry_price") or p.get("entry") or p.get("price") or 0
        stop = p.get("stop_loss", p.get("stop", 0)) or 0
        target = p.get("atr_target", p.get("target", 0)) or 0
        risk = entry_price - stop if (entry_price and stop) else 0
        pnl_share = res.get("pnl_dollar")
        r_realized = None
        if pnl_share is not None and risk and risk > 0:
            try:
                r_realized = round(float(pnl_share) / float(risk), 4)
            except (TypeError, ValueError, ZeroDivisionError):
                r_realized = None

        bt = _batch_time_24h(p.get("found_time", ""))
        cat = assign_category(score, label)
        cat_counts[cat] = cat_counts.get(cat, 0) + 1

        entries.append(normalize_entry({
            "date": date_str,
            "batch_time": bt,
            "ticker": p.get("ticker", ""),
            "entry": entry_price,
            "stop": stop,
            "target": target,
            "score": score,
            "rsi": p.get("rsi") if isinstance(p.get("rsi"), (int, float)) else (
                (p.get("indicators") or {}).get("rsi") if isinstance(p.get("indicators"), dict) else None
            ),
            "rvol": p.get("rvol", 0),
            "leadership_label": label,
            "category": cat,
            "result": res.get("result", ""),
            "resolve_time": res.get("resolve_time", ""),
            "resolve_price": res.get("exit_price"),
            "pnl_dollar": pnl_share,
            "r_realized": r_realized,
            "appearance": p.get("appearance_num", 1),
            "post_close": False,  # POSTCLOSE picks are already filtered upstream
            "note": p.get("note", ""),
        }))

    upsert_day(date_str, entries)
    logger.info(
        f"Performance log upserted for {date_str}: {len(entries)} entries "
        f"(A={cat_counts['A']} B={cat_counts['B']} C={cat_counts['C']} D={cat_counts['D']})"
    )


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Analyze daily scanner picks and append to Excel")
    parser.add_argument("date", nargs="?", default=None,
                        help="Date to analyze (YYYY-MM-DD). Default: today")
    parser.add_argument("--from-json", dest="json_path", default=None,
                        help="Path to a JSON file with picks (instead of fetching)")
    args = parser.parse_args()

    if args.date:
        target_date = args.date
    else:
        target_date = datetime.now(ET).strftime("%Y-%m-%d")

    analyze_day(target_date, args.json_path)
